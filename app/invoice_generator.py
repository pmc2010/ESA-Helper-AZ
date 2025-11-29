"""
Invoice Generator Module

Generates Excel and PDF invoices from template using vendor and student profiles.
"""

import logging
import json
import subprocess
import zipfile
import shutil
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

logger = logging.getLogger(__name__)

# Path to template - relative to this file's location
TEMPLATE_PATH = Path(__file__).parent / "templates" / "invoice_template.xltx"
# Define safe base directory for invoice output
BASE_OUTPUT_DIR = Path(__file__).parent.parent / "data" / "invoices"


def sanitize_filename(name: str, extension: str = None) -> str:
    """
    Sanitize input for use in a filename.
    Only allows alphanumerics, underscores, and hyphens (strict).
    Prevents path traversal, dotfile attacks, and command-line option injection.

    Args:
        name: Input filename to sanitize
        extension: Optional file extension to enforce (e.g., "xlsx", "pdf")

    Returns:
        Safe filename or raises ValueError if input is invalid

    Raises:
        ValueError: If filename cannot be safely sanitized
    """
    if not name:
        return "file"

    sanitized = str(name).strip()

    # Strict whitelist: only alphanumerics, underscores, hyphens
    # This rejects all special characters, spaces, dots, etc.
    if not re.fullmatch(r'[a-zA-Z0-9_\-]+', sanitized):
        # Fall back to replacing unsafe characters with underscore
        sanitized = re.sub(r'[^a-zA-Z0-9_\-]', '_', sanitized)

    # Remove leading dots, dashes, slashes - prevent dotfiles and option confusion
    sanitized = sanitized.lstrip('._-/')

    # Reject if still starts with problematic characters
    while sanitized and sanitized[0] in '.-/':
        sanitized = sanitized[1:]

    # Ensure not empty
    if not sanitized:
        sanitized = "file"

    # Collapse consecutive hyphens
    sanitized = re.sub(r'-+', '-', sanitized)

    # Optionally enforce a file extension
    if extension:
        # Remove any existing extension first
        sanitized = re.sub(r'\.[a-zA-Z0-9]+$', '', sanitized)
        # Add the enforced extension
        ext_clean = extension.lstrip('.')
        if not re.fullmatch(r'[a-zA-Z0-9]+', ext_clean):
            raise ValueError(f"Invalid extension: {extension}")
        sanitized = f"{sanitized}.{ext_clean}"

    return sanitized


class InvoiceGenerator:
    """Generate invoices from template with vendor and student data"""

    def __init__(self):
        """Initialize invoice generator"""
        if not TEMPLATE_PATH.exists():
            raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")
        self.template_path = TEMPLATE_PATH

    def _load_template(self):
        """Load template workbook with compatibility settings"""
        try:
            # Load with data_only=False to preserve formulas
            # and keep_vba=False to avoid template-specific issues
            wb = load_workbook(self.template_path)
            logger.info("✓ Template loaded with standard settings")
            return wb
        except Exception as e:
            logger.error(f"Error loading template: {str(e)}")
            raise

    def generate_invoice(self, vendor_data: dict, student_data: dict,
                        invoice_data: dict, output_dir: str) -> dict:
        """
        Generate invoice Excel and PDF files

        Args:
            vendor_data: Vendor profile {name, business_name, address_line_1, address_line_2, phone, email, tax_rate}
            student_data: Student profile {name, address_line_1, address_line_2}
            invoice_data: Invoice details {
                invoice_number: str,
                date: str (YYYY-MM-DD),
                line_items: list of {description: str, quantity: float, unit_price: float},
                tax_rate: float (optional, overrides vendor tax_rate)
                [DEPRECATED] description, quantity, unit_price - single item (will be converted to line_items)
            }
            output_dir: Directory to save files

        Returns:
            dict: {
                'success': bool,
                'excel_path': str,
                'pdf_path': str,
                'message': str
            }
        """
        try:
            logger.info("=" * 60)
            logger.info("INVOICE GENERATION")
            logger.info("=" * 60)

            # Validate output directory
            output_path = Path(output_dir)
            output_path.mkdir(parents=True, exist_ok=True)
            logger.info(f"Output directory: {output_path}")

            # Load template
            logger.info("1. Loading invoice template...")
            wb = self._load_template()
            ws = wb.active
            logger.info("✓ Template active worksheet loaded")

            # Fill vendor info
            logger.info("2. Filling vendor information...")
            # Use business_name if available, otherwise use vendor name
            vendor_display_name = vendor_data.get('business_name', '') or vendor_data.get('name', '')
            ws['B2'] = vendor_display_name
            ws['B3'] = vendor_data.get('address_line_1', '')
            ws['B4'] = vendor_data.get('address_line_2', '')
            ws['B5'] = vendor_data.get('phone', '')
            ws['B6'] = vendor_data.get('email', '')
            logger.info(f"✓ Vendor: {vendor_display_name}")

            # Fill student info (Bill To)
            logger.info("3. Filling student information...")
            ws['B9'] = student_data.get('name', '')
            ws['B10'] = student_data.get('address_line_1', '')
            ws['B11'] = student_data.get('address_line_2', '')
            logger.info(f"✓ Student: {student_data.get('name', 'N/A')}")

            # Fill invoice details
            logger.info("4. Filling invoice details...")
            ws['E9'] = invoice_data.get('invoice_number', '')
            ws['F9'] = invoice_data.get('date', datetime.now().strftime('%m/%d/%y'))
            logger.info(f"✓ Invoice #: {invoice_data.get('invoice_number', 'N/A')}")
            logger.info(f"✓ Date: {ws['F9'].value}")

            # Fill line items
            logger.info("5. Filling line items...")

            # Support both new format (line_items list) and old format (single description/quantity/unit_price)
            line_items = invoice_data.get('line_items', [])
            if not line_items and 'description' in invoice_data:
                # Backward compatibility: convert single item to line_items format
                line_items = [{
                    'description': invoice_data.get('description', ''),
                    'quantity': invoice_data.get('quantity', 1),
                    'unit_price': invoice_data.get('unit_price', 0)
                }]

            # Fill each line item starting at row 14
            start_row = 14
            for idx, item in enumerate(line_items):
                row = start_row + idx
                ws[f'B{row}'] = item.get('description', '')
                ws[f'D{row}'] = item.get('quantity', 1)
                ws[f'E{row}'] = item.get('unit_price', 0)
                # F{row} already has formula for line total - don't overwrite it
                logger.info(f"✓ Line {idx + 1}: {item.get('description', 'N/A')} - Qty: {ws[f'D{row}'].value}, Price: ${ws[f'E{row}'].value}")

            if line_items:
                logger.info(f"✓ Total line items: {len(line_items)}")

            # Set tax rate
            logger.info("6. Setting tax rate...")
            tax_rate = invoice_data.get('tax_rate', vendor_data.get('tax_rate', 0.0))
            ws['F30'] = tax_rate
            logger.info(f"✓ Tax Rate: {tax_rate * 100}%")

            # Save Excel file
            logger.info("7. Saving Excel file...")
            # Create filename with timestamp and vendor name
            # Use strict sanitization with enforced .xlsx extension
            invoice_number = sanitize_filename(str(invoice_data.get('invoice_number', 'invoice')))
            vendor_name = sanitize_filename(str(vendor_data.get('name', 'vendor')).lower())
            # Sanitize the combined filename and enforce extension
            excel_filename = sanitize_filename(f"{invoice_number}_{vendor_name}_inv", extension="xlsx")
            excel_path = output_path / excel_filename

            # Validate that excel_path stays within BASE_OUTPUT_DIR (prevent path traversal)
            self._validate_output_path(excel_path, BASE_OUTPUT_DIR)

            # Save Excel file - preserve template structure to avoid corruption
            try:
                logger.info("7. Saving Excel file...")
                self._save_invoice_preserving_template(wb, excel_path)
                logger.info(f"✓ Saved: {excel_path}")
            except Exception as e:
                logger.error(f"Excel save failed: {str(e)}")
                raise

            # Verify file was created and has content
            if not excel_path.exists():
                raise FileNotFoundError(f"Failed to create Excel file: {excel_path}")
            file_size = excel_path.stat().st_size
            if file_size < 5000:  # Warn if file seems too small (might be corrupted)
                logger.warning(f"⚠ File created but seems small: {file_size} bytes - may be corrupted")
            else:
                logger.info(f"✓ File created: {file_size} bytes")

            # Recalculate formulas in Excel file using LibreOffice
            logger.info("7b. Recalculating formulas...")
            self._recalculate_excel_formulas(excel_path)

            # Convert to PDF
            logger.info("8. Converting to PDF...")
            # Sanitize with enforced .pdf extension
            pdf_filename = sanitize_filename(f"{invoice_number}_{vendor_name}_inv", extension="pdf")
            pdf_path = output_path / pdf_filename

            # Validate that pdf_path stays within BASE_OUTPUT_DIR (prevent path traversal)
            self._validate_output_path(pdf_path, BASE_OUTPUT_DIR)

            if self._convert_to_pdf(excel_path, pdf_path):
                logger.info(f"✓ Saved: {pdf_path}")

                # Clean up Excel file after PDF creation (optional)
                # Uncomment if you only want to keep PDF
                # excel_path.unlink()

                return {
                    'success': True,
                    'excel_path': str(excel_path),
                    'pdf_path': str(pdf_path),
                    'message': 'Invoice generated successfully'
                }
            else:
                logger.warning("PDF conversion failed, but Excel file created")
                return {
                    'success': True,  # Partial success - Excel created even if PDF fails
                    'excel_path': str(excel_path),
                    'pdf_path': None,
                    'message': 'Invoice generated (Excel only - PDF conversion failed)'
                }

        except Exception as e:
            logger.error(f"❌ Error generating invoice: {str(e)}")
            logger.error(f"Full traceback:", exc_info=True)
            return {
                'success': False,
                'excel_path': None,
                'pdf_path': None,
                'message': f'Error: {str(e)}'
            }

    def _save_invoice_preserving_template(self, wb, excel_path: Path):
        """
        Save invoice while preserving template structure (custom XML, metadata, etc.)

        The template file (.xltx) contains custom XML and metadata that openpyxl normally
        strips when saving. This method preserves those by:
        1. Copying the original template file to the output path
        2. Extracting the modified worksheet XML from openpyxl
        3. Injecting it into the copied template file

        Args:
            wb: Modified workbook from openpyxl
            excel_path: Path where to save the file
        """
        try:
            # Step 1: Save the modified workbook to a temporary file
            temp_path = excel_path.parent / f".tmp_{excel_path.name}"
            logger.info(f"   Saving modified workbook to temp file...")
            wb.save(temp_path)

            # Step 2: Copy the original template file to the output location
            logger.info(f"   Copying template structure...")
            shutil.copy2(self.template_path, excel_path)

            # Step 3: Extract worksheet data from temp file and inject into copied template
            logger.info(f"   Injecting modified data into template...")
            self._merge_worksheet_into_template(temp_path, excel_path)

            # Step 4: Clean up temp file
            temp_path.unlink(missing_ok=True)

            logger.info(f"   ✓ Invoice saved while preserving template structure")

        except Exception as e:
            logger.error(f"   Error preserving template structure: {str(e)}")
            # Fallback: save normally but still fix content type
            logger.warning(f"   Falling back to standard save (with content type fix)...")
            temp_path = excel_path.parent / f".tmp_{excel_path.name}"
            wb.save(temp_path)
            # Still need to fix content type even for normal save
            self._fix_content_type_in_file(temp_path, excel_path)
            temp_path.unlink(missing_ok=True)

    def _merge_worksheet_into_template(self, temp_path: Path, template_path: Path):
        """
        Extract the modified worksheet from temp file and inject it into the template file.
        This preserves all custom XML and metadata from the original template.

        Args:
            temp_path: Path to temp file with modified workbook (created by openpyxl)
            template_path: Path to template file (now a copy of original)
        """
        try:
            # Extract the worksheet XML from the temp file
            worksheet_content = None
            styles_content = None
            with zipfile.ZipFile(temp_path, 'r') as temp_zip:
                # Get the worksheet content
                try:
                    worksheet_content = temp_zip.read('xl/worksheets/sheet1.xml')
                except KeyError:
                    pass
                # Get the styles
                try:
                    styles_content = temp_zip.read('xl/styles.xml')
                except KeyError:
                    pass

            if not worksheet_content:
                raise ValueError("Could not extract worksheet content from temp file")

            # Create a temporary output file for the merged zip
            temp_output = template_path.parent / f".tmp_merge_{template_path.name}"

            # Read from template, update worksheets, write to temp output
            with zipfile.ZipFile(template_path, 'r') as template_zip:
                with zipfile.ZipFile(temp_output, 'w', zipfile.ZIP_DEFLATED) as out_zip:
                    # Copy all files from template, except worksheet and optionally styles
                    for item in template_zip.infolist():
                        if item.filename == 'xl/worksheets/sheet1.xml':
                            # Replace with modified worksheet
                            out_zip.writestr(item, worksheet_content)
                        elif item.filename == 'xl/styles.xml' and styles_content:
                            # Replace with modified styles
                            out_zip.writestr(item, styles_content)
                        elif item.filename == '[Content_Types].xml':
                            # Fix content type: template.main -> sheet.main
                            # This is critical! Without this fix, Excel rejects the file
                            # because it thinks it's a template (.xltx) not a workbook (.xlsx)
                            content = template_zip.read(item.filename).decode('utf-8')
                            content = content.replace(
                                'spreadsheetml.template.main+xml',
                                'spreadsheetml.sheet.main+xml'
                            )
                            out_zip.writestr(item, content.encode('utf-8'))
                            logger.info(f"   ✓ Fixed content type (template.main -> sheet.main)")
                        else:
                            # Copy as-is
                            out_zip.writestr(item, template_zip.read(item.filename))

            # Replace original with merged file
            template_path.unlink()
            temp_output.rename(template_path)

            logger.info(f"   ✓ Worksheet injected into template")

        except Exception as e:
            logger.error(f"   Failed to merge worksheet: {str(e)}")
            raise

    def _fix_content_type_in_file(self, temp_path: Path, output_path: Path):
        """
        Fix the [Content_Types].xml to use worksheet content type instead of template.

        The template file uses 'spreadsheetml.template.main+xml' but we need
        'spreadsheetml.sheet.main+xml' for a regular Excel file that Excel can open.

        Args:
            temp_path: Path to temp file with wrong content type
            output_path: Path to final output file
        """
        try:
            # Create a temporary output file for the corrected zip
            temp_output = output_path.parent / f".tmp_ct_{output_path.name}"

            # Read the temp file and correct the content type
            with zipfile.ZipFile(temp_path, 'r') as temp_zip:
                with zipfile.ZipFile(temp_output, 'w', zipfile.ZIP_DEFLATED) as out_zip:
                    for item in temp_zip.infolist():
                        if item.filename == '[Content_Types].xml':
                            # Read and fix the content type
                            content = temp_zip.read(item.filename).decode('utf-8')
                            # Replace template.main with sheet.main
                            if 'template.main' in content:
                                content = content.replace(
                                    'spreadsheetml.template.main+xml',
                                    'spreadsheetml.sheet.main+xml'
                                )
                                out_zip.writestr(item, content.encode('utf-8'))
                                logger.info(f"   ✓ Fixed content type (template.main -> sheet.main)")
                            else:
                                # Already correct, copy as-is
                                out_zip.writestr(item, temp_zip.read(item.filename))
                        else:
                            # Copy everything else as-is
                            out_zip.writestr(item, temp_zip.read(item.filename))

            # Replace original with corrected file
            temp_output.rename(output_path)

        except Exception as e:
            logger.error(f"   Failed to fix content type: {str(e)}")
            raise

    def _recalculate_excel_formulas(self, excel_path: Path):
        """
        Recalculate Excel formulas using LibreOffice in-place

        Excel files created by openpyxl contain formulas but Excel doesn't calculate
        them until the file is opened and saved. This method uses LibreOffice to
        open the file, recalculate, and save it back.

        Args:
            excel_path: Path to Excel file to recalculate
        """
        try:
            # Check if LibreOffice is available
            result = subprocess.run(['which', 'libreoffice'],
                                  capture_output=True, text=True)
            if result.returncode != 0:
                logger.warning("   LibreOffice not found - formulas will not be calculated")
                logger.warning("   Install with: brew install libreoffice")
                return

            logger.info(f"   Opening Excel file with LibreOffice for formula recalculation...")

            # Create a Python script that LibreOffice will execute to recalculate
            # This uses LibreOffice's UNO bridge to properly calculate formulas
            # Use sanitized filenames for temp files
            safe_excel_name = sanitize_filename(excel_path.stem)
            temp_script = excel_path.parent / f".tmp_{safe_excel_name}_recalc.py"
            temp_output = excel_path.parent / f".tmp_{safe_excel_name}"

            # Validate temp script and output paths are within BASE_OUTPUT_DIR (prevent path traversal)
            self._validate_output_path(temp_script, BASE_OUTPUT_DIR)
            self._validate_output_path(temp_output, BASE_OUTPUT_DIR)

            script_content = f'''
import uno
from com.sun.star.beans import PropertyValue

def recalc():
    try:
        # Open document
        local_context = uno.getComponentContext()
        resolver = local_context.ServiceManager.createInstanceWithContext(
            "com.sun.star.bridge.UnoUrlResolver", local_context)
        try:
            ctx = resolver.resolve(
                "uno:socket,host=localhost,port=2002;urp;StarOffice.ComponentContext")
            smgr = ctx.ServiceManager
        except:
            # Start fresh LibreOffice instance
            import subprocess
            subprocess.Popen(['libreoffice', '--headless', '--accept=socket,host=localhost,port=2002;urp;'])
            import time
            time.sleep(3)
            ctx = resolver.resolve(
                "uno:socket,host=localhost,port=2002;urp;StarOffice.ComponentContext")
            smgr = ctx.ServiceManager

        desktop = smgr.createInstanceWithContext("com.sun.star.frame.Desktop", ctx)
        doc = desktop.loadComponentFromURL("file://{excel_path}", "_blank", 0, ())

        # Recalculate
        doc.calculateAll()

        # Save
        doc.store()
        doc.close(True)

    except Exception as e:
        print(f"Error: {{e}}")

g_exportedScripts = (recalc,)
'''

            try:
                # Write script - validate path safety one more time before write
                # Use canonical path check with .relative_to() only (no fragile string prefix checks)
                script_path_real = temp_script.resolve()
                base_dir_real = BASE_OUTPUT_DIR.resolve()
                try:
                    script_path_real.relative_to(base_dir_real)
                except ValueError:
                    raise ValueError(f"Attempted to write temp script outside of safe directory: {script_path_real}")

                with open(temp_script, 'w') as f:
                    f.write(script_content)

                # Validate excel_path before passing to subprocess
                # Check: exists, not a symlink, within BASE_OUTPUT_DIR, safe filename
                excel_path_real = excel_path.resolve()
                base_dir_real = BASE_OUTPUT_DIR.resolve()

                # Ensure it's within safe directory
                try:
                    excel_path_real.relative_to(base_dir_real)
                except ValueError:
                    raise ValueError(f"Excel path is not within safe output directory: {excel_path_real}")

                # Reject symlinks to prevent symlink attacks
                if excel_path_real.is_symlink():
                    raise ValueError(f"Excel file is a symlink, refusing to process: {excel_path_real}")

                # Check file exists
                if not excel_path_real.exists():
                    raise ValueError(f"Excel file does not exist: {excel_path_real}")

                # Validate filename matches safe pattern (alphanumerics, underscore, hyphen, dot, .xlsx extension)
                safe_filename_pattern = re.compile(r'^[a-zA-Z0-9_\-]+\.xlsx$')
                if not safe_filename_pattern.match(excel_path_real.name):
                    raise ValueError(f"Unsafe excel filename: {excel_path_real.name}")

                # Try UNO approach first
                cmd = [
                    'libreoffice',
                    '--headless',
                    f'--script-provider=python:{temp_script}',
                    str(excel_path_real)
                ]
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)

                if result.returncode != 0:
                    # Fallback: Simple open and save approach
                    logger.info("   Using fallback recalculation method...")
                    cmd = [
                        'libreoffice',
                        '--headless',
                        '--calc',
                        '--invisible',
                        str(excel_path_real)
                    ]
                    result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)

                logger.info("   ✓ Formulas recalculated")

            finally:
                # Clean up script - validate before deleting
                try:
                    # Validate path is safe to delete (within BASE_OUTPUT_DIR)
                    self._validate_output_path(temp_script, BASE_OUTPUT_DIR)
                    temp_script.unlink()
                except ValueError as ve:
                    logger.error(f"Security validation failed before temp script cleanup: {ve}")
                except:
                    pass

        except subprocess.TimeoutExpired:
            logger.warning("   LibreOffice recalculation timed out")
        except FileNotFoundError:
            logger.warning("   LibreOffice not found - formulas will not be calculated")
            logger.warning("   Install with: brew install libreoffice")
        except Exception as e:
            logger.warning(f"   Failed to recalculate formulas: {str(e)}")
            # Don't raise - Excel file is still usable, just without calculated values

    def _convert_to_pdf(self, excel_path: Path, pdf_path: Path) -> bool:
        """
        Convert Excel file to PDF using LibreOffice

        Args:
            excel_path: Path to Excel file
            pdf_path: Path to save PDF

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            logger.info(f"   Converting {excel_path.name} to PDF...")

            # Try to find LibreOffice executable
            # macOS native installation
            libreoffice_path = '/Applications/LibreOffice.app/Contents/MacOS/soffice'
            # Homebrew installations (various paths)
            homebrew_paths = [
                '/usr/local/bin/libreoffice',
                '/opt/homebrew/bin/libreoffice',
                'libreoffice'  # fallback to PATH
            ]

            libreoffice_cmd = None
            if Path(libreoffice_path).exists():
                libreoffice_cmd = libreoffice_path
                logger.info(f"   Using macOS LibreOffice: {libreoffice_path}")
            else:
                for path in homebrew_paths:
                    if path == 'libreoffice':
                        # Try PATH lookup
                        try:
                            result = subprocess.run(['which', 'libreoffice'],
                                                   capture_output=True, text=True)
                            if result.returncode == 0:
                                libreoffice_cmd = result.stdout.strip()
                                logger.info(f"   Found LibreOffice in PATH: {libreoffice_cmd}")
                                break
                        except:
                            pass
                    elif Path(path).exists():
                        libreoffice_cmd = path
                        logger.info(f"   Found LibreOffice: {path}")
                        break

            if not libreoffice_cmd:
                logger.error("LibreOffice not found in common locations")
                logger.error("   macOS: /Applications/LibreOffice.app (or brew install libreoffice)")
                logger.error("   Linux: sudo apt-get install libreoffice")
                return False

            # Use LibreOffice headless conversion
            result = subprocess.run([
                libreoffice_cmd,
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', str(pdf_path.parent),
                str(excel_path)
            ], capture_output=True, timeout=30)

            if result.returncode == 0:
                # LibreOffice creates PDF with .pdf extension based on input filename
                # Rename if necessary to match desired name
                generated_pdf = pdf_path.parent / (excel_path.stem + '.pdf')
                if generated_pdf.exists() and generated_pdf != pdf_path:
                    generated_pdf.rename(pdf_path)
                logger.info(f"   ✓ PDF conversion successful")
                return True
            else:
                error_msg = result.stderr.decode() if result.stderr else "Unknown error"
                logger.error(f"   LibreOffice error: {error_msg}")
                return False

        except subprocess.TimeoutExpired:
            logger.error("   LibreOffice conversion timed out (30 seconds)")
            return False
        except Exception as e:
            logger.error(f"   PDF conversion error: {str(e)}")
            return False

    def _validate_output_path(self, path: Path, base_dir: Path) -> None:
        """
        Validate that path is strictly contained within base_dir.
        Prevents path traversal attacks.

        Args:
            path: The path to validate
            base_dir: The base directory that path must be within

        Raises:
            ValueError: If path escapes base_dir
        """
        try:
            base_dir_resolved = base_dir.resolve()
            path_resolved = path.resolve()

            # Use is_relative_to (Python 3.9+) for robust containment check
            # This prevents false positives from startswith() checks
            try:
                path_resolved.relative_to(base_dir_resolved)
                logger.debug(f"Path validation passed: {path_resolved}")
            except ValueError:
                logger.error(f"Path validation failed: {path_resolved} is not within {base_dir_resolved}")
                raise ValueError(f"Unsafe path: {path} must be within {base_dir}")

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Path validation error: {str(e)}")
            raise ValueError(f"Failed to validate path {path}: {str(e)}")


def load_vendor_profiles() -> list:
    """Load vendor profiles from JSON file"""
    try:
        vendors_file = Path(__file__).parent.parent / 'data' / 'vendors.json'
        with open(vendors_file, 'r') as f:
            data = json.load(f)
            return data.get('vendors', [])
    except Exception as e:
        logger.error(f"Error loading vendors: {str(e)}")
        return []


def load_student_profiles() -> list:
    """Load student profiles from JSON file"""
    try:
        students_file = Path(__file__).parent.parent / 'data' / 'students.json'
        with open(students_file, 'r') as f:
            data = json.load(f)
            return data.get('students', [])
    except Exception as e:
        logger.error(f"Error loading students: {str(e)}")
        return []


def get_vendor(vendor_name: str) -> dict:
    """Get vendor profile by name (supports partial/first name matching)"""
    profiles = load_vendor_profiles()

    # First try exact match (e.g., "Math Tutoring Services")
    exact_match = next((v for v in profiles if v['name'].lower() == vendor_name.lower()), None)
    if exact_match:
        return exact_match

    # Try matching by first name (e.g., "Math" matches "Math Tutoring Services")
    # or if the profile name starts with the provided name
    partial_match = next((v for v in profiles
                         if v['name'].lower().split()[0] == vendor_name.lower() or
                         v['name'].lower().startswith(vendor_name.lower())), None)
    if partial_match:
        return partial_match

    # No match found
    return None


def get_student(student_name: str) -> dict:
    """Get student profile by name (supports partial/first name matching)"""
    profiles = load_student_profiles()

    # First try exact match (e.g., "Student One")
    exact_match = next((s for s in profiles if s['name'].lower() == student_name.lower()), None)
    if exact_match:
        return exact_match

    # Try matching by first name (e.g., "Student" matches "Student One")
    # or if the profile name starts with the provided name
    partial_match = next((s for s in profiles
                         if s['name'].lower().split()[0] == student_name.lower() or
                         s['name'].lower().startswith(student_name.lower())), None)
    if partial_match:
        return partial_match

    # No match found
    return None


def save_vendor_profile(vendor: dict) -> bool:
    """Save vendor profile to JSON file"""
    try:
        vendors_file = Path(__file__).parent.parent / 'data' / 'vendors.json'
        profiles = load_vendor_profiles()

        # Update or add vendor
        existing = next((v for v in profiles if v['id'] == vendor['id']), None)
        if existing:
            profiles.remove(existing)
        profiles.append(vendor)

        with open(vendors_file, 'w') as f:
            json.dump({'vendors': profiles}, f, indent=2)

        logger.info(f"✓ Vendor profile saved: {vendor['name']}")
        return True
    except Exception as e:
        logger.error(f"Error saving vendor: {str(e)}")
        return False


def save_student_profile(student: dict) -> bool:
    """Save student profile to JSON file"""
    try:
        students_file = Path(__file__).parent.parent / 'data' / 'students.json'
        profiles = load_student_profiles()

        # Update or add student
        existing = next((s for s in profiles if s['id'] == student['id']), None)
        if existing:
            profiles.remove(existing)
        profiles.append(student)

        with open(students_file, 'w') as f:
            json.dump({'students': profiles}, f, indent=2)

        logger.info(f"✓ Student profile saved: {student['name']}")
        return True
    except Exception as e:
        logger.error(f"Error saving student: {str(e)}")
        return False
